import openpyxl
from openpyxl.styles import Font #字体格式和大小
from openpyxl.styles import Alignment #单元格高度和宽度

myxlsx = '练习文件/UCM_USER.xlsx'
wb = openpyxl.load_workbook(myxlsx,data_only=True)
print(type(wb))

allSheets = wb.get_sheet_names()
print('所有的sheets页名称=',allSheets)

ws = wb.get_sheet_by_name('女的')
print('当前工作表=',ws.title)
print('单元格1 = ',ws['A2'].value)
print('单元格2 = ',ws['B2'].value)
print('单元格3 = ',ws['C2'].value)
print('单元格4 = ',ws['D2'].value)
print('总共 %d 列'%ws.max_column)
print('总共 %d 行'%ws.max_row)
for i in range(4,7):
    for j in range(1,6):
        print(ws.cell(column=i,row=j).value,end=' ')
    print()    
#print(ws['AS'].value)

for row in ws.rows:
    for cell in row:
        print(cell.value,end=' ')
    print()  

wb1=openpyxl.Workbook()
ws1=wb1.active

print('目前工作表名称：',ws1.title)
ws1.title = 'My sheet'
print('新工作表名称:',ws1.title)
wb1.save('My sheet.xlsx')
wb1.create_sheet()
print('创建sheet成功')
wb1.create_sheet(index=0,title='sheet1')
print('sheet1成功')
wb1.create_sheet(index=2,title='sheet2')
print('sheet2成功')
wb1.save('My sheet.xlsx')
wb1.remove_sheet(wb1.get_sheet_by_name('sheet1'))
print('删除sheet1成功')
wb1.save('My sheet.xlsx')
ws1['A1'] = '姓名'
ws1['B1'] = '刘安珂'
print('写入单元格数据成功')
row1 = ['姓名','性别','年龄']
ws1.append(row1)
row2 = ['刘安珂','女','26']
ws1.append(row2)
print('将列表数据写进单元格')
wb1.save('My sheet.xlsx')
fonttitle = Font(name='宋体',size=25,color='FF66FF')
ws1['A1'].font= fonttitle #字体格式
ws1['A1'].alignment = Alignment(horizontal='center',vertical='center')#居中方式
ws1.merge_cells('A1:E1') #合并单元格
ws1.unmerge_cells('A1:E1') #合并单元格
ws1['A1'] = '名称详情'
ws1.row_dimensions[1].height =30 #高度
ws1.row_dimensions[1].width =30 #宽度
row = [
    ['姓名','性别','年龄'],
    ['刘安珂','女',26],
    ['金晶','女',28],
    ['刘亦菲','女',35]
    ]
for name in row:
    ws1.append(name)
print('循环将列表数据进行保存')
wb1.save('My sheet.xlsx')

#图表练习-柱状图
from openpyxl.chart import BarChart,Reference#图形包

wb2=openpyxl.Workbook()
ws2=wb2.active
row = [
    ['姓名','性别','年龄'],
    ['刘安珂','女',26],
    ['金晶','女',28],
    ['刘亦菲','女',35]
    ]
for name in row:
    ws2.append(name)
print('插入数据成功')    

chart =BarChart()
chart.title='人名详情'
chart.y_axis.title='姓名'
chart.x_axis.title='年龄'
#min_col,max_col这两个参数是列
#min_row,max_row这两个参数是行
data=Reference(ws2,min_col=3,max_col=3,min_row=1,max_row=5)
chart.add_data(data,titles_from_data=True)
xtitle=Reference(ws2,min_col=1,min_row=2,max_row=5)
chart.set_categories(xtitle)
ws2.add_chart(chart,'E1')
wb2.save('人名柱状图.xlsx')



#图表练习-饼图
from openpyxl.chart import PieChart,Reference#图形包

wb2=openpyxl.Workbook()
ws2=wb2.active
row = [
    ['姓名','性别','年龄'],
    ['刘安珂','女',26],
    ['金晶','女',28],
    ['刘亦菲','女',35]
    ]
for name in row:
    ws2.append(name)
print('插入数据成功')    

chart =PieChart()
chart.title='人名详情'
#min_col,max_col这两个参数是列
#min_row,max_row这两个参数是行
data=Reference(ws2,min_col=3,max_col=3,min_row=1,max_row=5)
chart.add_data(data,titles_from_data=True)
xtitle=Reference(ws2,min_col=1,min_row=2,max_row=5)
chart.set_categories(xtitle)
ws2.add_chart(chart,'E1')
wb2.save('人名饼图.xlsx')




